import openpyxl

book = openpyxl.load_workbook("C:\\Users\\HI\\OneDrive\\Documents\\data.xlsx")

sheet = book.active

cell = sheet.cell(row=1, column=2)

print(cell.value)

cell1 = sheet.cell(row=2, column=3)
cell1.value = "kalasala"

print(cell1.value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A4'].value)

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)


